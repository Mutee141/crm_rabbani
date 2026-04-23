from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.utils import timezone
from django.db.models import Q, Sum, Count, F
from datetime import datetime
import openpyxl
from django.http import HttpResponse
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.db.models import Count, Q
# Import your choices and models
from .models import Employee, AttendanceRecord, AttendanceSchedule, AttendanceUpload, DEPARTMENT_CHOICES, DESIGNATION_CHOICES

from .models import (
    Employee,
    Leave,
    AttendanceUpload,
    AttendanceSchedule,
    AttendanceRecord,
    DEPARTMENT_CHOICES,
    DESIGNATION_CHOICES
)
from .forms import (
    EmployeeForm,
    LeaveForm,
    AttendanceUploadForm,
    AttendanceScheduleForm,
    AttendanceRecordForm,
)


from users.decorators import role_required

@role_required("HR")
def hr_dashboard(request):
    return render(request, "hr/dashboard.html")




# -------------------------
# Dashboard Page
# -------------------------
from django.utils import timezone
from datetime import timedelta
from django.db.models import Count, Q
import json

@role_required("HR")
def hr_dashboard(request):
    today = timezone.now().date()
    
    # 1. Top Row Stats
    total_employees = Employee.objects.filter(status="Active").count()
    
    # Get attendance for today
    attendance_today = AttendanceRecord.objects.filter(date=today)
    
    present_today = attendance_today.filter(status="Present").count()
    absent_today = attendance_today.filter(status="Absent").count()
    full_leave_today = attendance_today.filter(status="Full Leave").count()
    short_leave_today = attendance_today.filter(status="Short Leave").count()

    # 2. Graph Logic: Last 7 Days Attendance Trend
    last_7_days = []
    days_labels = []
    present_counts = []
    absent_counts = []

    for i in range(6, -1, -1):
        day = today - timedelta(days=i)
        days_labels.append(day.strftime('%b %d'))
        
        stats = AttendanceRecord.objects.filter(date=day).aggregate(
            p=Count('id', filter=Q(status="Present")),
            a=Count('id', filter=Q(status="Absent"))
        )
        present_counts.append(stats['p'])
        absent_counts.append(stats['a'])

    # 3. Check-in/Check-out Details (Day by Day Table)
    # Showing last 30 logs ordered by date
    daily_logs = AttendanceRecord.objects.select_related('employee').order_by("-date", "-check_in")[:30]

    context = {
        "total_employees": total_employees,
        "present_today": present_today,
        "absent_today": absent_today,
        "full_leave_today": full_leave_today,
        "short_leave_today": short_leave_today,
        
        # Chart Data
        "days_labels": json.dumps(days_labels),
        "present_counts": json.dumps(present_counts),
        "absent_counts": json.dumps(absent_counts),
        
        # For Today's Stats Chart
        "today_stats_data": json.dumps([present_today, absent_today, full_leave_today, short_leave_today]),
        
        "daily_logs": daily_logs,
    }
    return render(request, "hr/dashboard.html", context)


# -------------------------
# EMPLOYEE CRUD
# -------------------------
# hr/views.py

def employee_list(request):
    # Fetch all employees ordered by latest
    employees = Employee.objects.all().order_by('-created_at')
    
    # Get filters from GET request
    search_name = request.GET.get('search', '').strip()
    filter_department = request.GET.get('department', '').strip()
    
    # Apply filters
    if search_name:
        employees = employees.filter(
            Q(name__icontains=search_name) | Q(emp_id__icontains=search_name)
        )
    
    if filter_department and filter_department != 'all':
        employees = employees.filter(department=filter_department)
    
    # Calculate stats
    total_employees = Employee.objects.all().count()
    active_employees = Employee.objects.filter(status='Active').count()
    total_departments = Employee.objects.values('department').distinct().count()
    
    # Get unique departments for dropdown
    departments = [choice[0] for choice in DEPARTMENT_CHOICES]
    
    # Check if a form was submitted from the Modal
    if request.method == "POST":
        form = EmployeeForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, "Employee registered successfully!")
            return redirect("hr:employee_list")
        else:
            messages.error(request, "Registration failed. Please check the form errors.")
    else:
        form = EmployeeForm()

    context = {
        "employees": employees,
        "form": form,
        "total_employees": total_employees,
        "active_employees": active_employees,
        "total_departments": total_departments,
        "departments": departments,
        "search_name": search_name,
        "filter_department": filter_department,
    }
    return render(request, "hr/employee/list.html", context)


def employee_create(request):
    if request.method == "POST":
        form = EmployeeForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, "Employee added successfully!")
            return redirect("hr:employee_list")
    else:
        form = EmployeeForm()

    return render(request, "hr/employee/create.html", {"form": form})


def employee_update(request, pk):
    emp = get_object_or_404(Employee, pk=pk)

    if request.method == "POST":
        form = EmployeeForm(request.POST, request.FILES, instance=emp)
        if form.is_valid():
            form.save()
            messages.success(request, f"Changes for {emp.name} saved successfully!")
            # CHANGE THIS LINE BELOW:
            return redirect("hr:employee_list") 
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = EmployeeForm(instance=emp)

    return render(request, "hr/employee/edit.html", {"form": form, "emp": emp})


def employee_delete(request, pk):
    emp = get_object_or_404(Employee, pk=pk)
    emp.delete()
    messages.success(request, "Employee deleted!")
    return redirect("hr:employee_list")


# -------------------------
# LEAVE CRUD
# -------------------------
def leave_list(request):
    leaves = Leave.objects.select_related('employee').all().order_by('-created_at')
    
    # Calculate stats
    total_leaves = leaves.count()
    full_leaves = leaves.filter(leave_type='Full Leave').count()
    short_leaves = leaves.filter(leave_type='Short Leave').count()
    on_duty = leaves.filter(leave_type='On Duty').count()
    
    return render(request, "hr/leave/list.html", {
        "leaves": leaves,
        "total_leaves": total_leaves,
        "full_leaves": full_leaves,
        "short_leaves": short_leaves,
        "on_duty": on_duty,
    })

def leave_create(request):
    if request.method == "POST":
        form = LeaveForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, "Leave created successfully!")
            return redirect("hr:leave_list") # Fixed: added hr:
    else:
        form = LeaveForm()
    return render(request, "hr/leave/create.html", {"form": form})

def leave_update(request, pk):
    leave = get_object_or_404(Leave, pk=pk)
    if request.method == 'POST':
        form = LeaveForm(request.POST, request.FILES, instance=leave)
        if form.is_valid():
            form.save()
            messages.success(request, "Leave updated successfully.")
            return redirect('hr:leave_list') # Fixed: added hr:
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = LeaveForm(instance=leave)
    
    return render(request, 'hr/leave/edit.html', {'form': form, 'leave': leave})

def leave_delete(request, pk):
    leave = get_object_or_404(Leave, pk=pk)
    if request.method == 'POST':
        employee_name = leave.employee.name 
        leave.delete()
        messages.success(request, f"Leave request for {employee_name} successfully deleted.")
        return redirect('hr:leave_list') # Fixed: added hr:
    
    return redirect('hr:leave_list') # Fixed: added hr:



# -------------------------------------------------------
# ------------------------- ATTENDANCE ------------------
# -------------------------------------------------------

# 1️⃣ Upload Attendance File
from django.shortcuts import render, redirect
from django.contrib import messages
from openpyxl import load_workbook
from django.db import transaction

from .forms import AttendanceUploadForm
from .models import AttendanceUpload

from datetime import datetime, timedelta

from datetime import datetime, time, date

from datetime import datetime, date, time
from django.db import transaction
from django.shortcuts import render, redirect
from django.contrib import messages
from openpyxl import load_workbook
from hr.models import Record



def normalize_time(value):
    if value is None:
        return None
    if isinstance(value, time):
        return value
    if isinstance(value, datetime):
        return value.time()
    if isinstance(value, str):
        try:
            return datetime.strptime(value.strip(), "%H:%M").time()
        except ValueError:
            return None
    return None

from datetime import datetime, date, timedelta, time
from django.shortcuts import render, redirect
from django.contrib import messages
from django.db import transaction
from openpyxl import load_workbook
# Ensure AttendanceRecord is imported correctly
from .models import AttendanceUpload, AttendanceRecord, AttendanceSchedule, Employee, Leave
from .forms import AttendanceUploadForm

def normalize_time(time_val):
    """Helper to handle various Excel time formats (float, string, or time object)"""
    if isinstance(time_val, time):
        return time_val
    if isinstance(time_val, (int, float)):
        # Handle Excel fractional day format (e.g., 0.375 for 9:00 AM)
        total_seconds = int(time_val * 86400)
        hours = total_seconds // 3600
        minutes = (total_seconds % 3600) // 60
        return time(hour=min(hours, 23), minute=min(minutes, 59))
    if isinstance(time_val, str) and ":" in time_val:
        try:
            return datetime.strptime(time_val.strip(), "%H:%M").time()
        except:
            return None
    return None

def attendance_upload(request):
    uploads = AttendanceUpload.objects.all().order_by("-uploaded_at")
    
    if request.method == "POST":
        form = AttendanceUploadForm(request.POST, request.FILES)
        if form.is_valid():
            upload = form.save(commit=False)
            upload.status = "Pending"
            upload.save()

            try:
                wb = load_workbook(upload.file.path, data_only=True)
                sheet = wb.active
                success_count = 0
                error_logs = []

                with transaction.atomic():
                    for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                        if not any(row): continue
                        
                        raw_emp_id, raw_date, raw_in, raw_out = row

                        # --- IMPROVED EMPLOYEE LOOKUP ---
                        try:
                            # Convert to string and strip decimals (e.g., 101.0 -> 101)
                            emp_id_str = str(raw_emp_id).split('.')[0].strip()
                            employee = Employee.objects.get(emp_id=emp_id_str)
                        except (Employee.DoesNotExist, ValueError):
                            error_logs.append(f"Row {index}: ID '{raw_emp_id}' not found.")
                            continue

                        # --- DATE NORMALIZATION ---
                        if isinstance(raw_date, datetime):
                            record_date = raw_date.date()
                        elif isinstance(raw_date, date):
                            record_date = raw_date
                        else:
                            try:
                                record_date = datetime.strptime(str(raw_date), "%Y-%m-%d").date()
                            except:
                                error_logs.append(f"Row {index}: Invalid date format.")
                                continue

                        # --- TIME NORMALIZATION ---
                        check_in = normalize_time(raw_in)
                        check_out = normalize_time(raw_out)

                        # --- SAVE TO AttendanceRecord ---
                        AttendanceRecord.objects.update_or_create(
                            employee=employee,
                            date=record_date,
                            defaults={
                                "check_in": check_in,
                                "check_out": check_out,
                                "status": "Present" if check_in else "Absent"
                            }
                        )
                        success_count += 1

                upload.status = "Completed"
                upload.record_count = success_count
                upload.remarks = " | ".join(error_logs[:3]) if error_logs else "Success"
                upload.save()

                if success_count > 0:
                    messages.success(request, f"Successfully processed {success_count} records.")
                else:
                    messages.warning(request, "Upload completed but 0 records were saved. Check Employee IDs.")
                
                return redirect("hr:attendance_upload")

            except Exception as e:
                upload.status = "Failed"
                upload.remarks = str(e)
                upload.save()
                messages.error(request, f"Upload failed: {e}")
                return redirect("hr:attendance_upload")
    else:
        form = AttendanceUploadForm()
    
    return render(request, "hr/attendance/upload.html", {"form": form, "uploads": uploads})

def generate_attendance(request):
    employees = Employee.objects.filter(status="Active").order_by("name")
    from_date_str = request.GET.get("from_date")
    to_date_str = request.GET.get("to_date")
    employee_id = request.GET.get("employee")

    results = []
    has_searched = False

    if from_date_str and to_date_str:
        has_searched = True
        from_date = datetime.strptime(from_date_str, "%Y-%m-%d").date()
        to_date = datetime.strptime(to_date_str, "%Y-%m-%d").date()

        employee_qs = employees
        if employee_id and employee_id != "all":
            employee_qs = employee_qs.filter(id=employee_id)

        current_date = from_date
        while current_date <= to_date:
            for employee in employee_qs:
                # Query AttendanceRecord
                record = AttendanceRecord.objects.filter(employee=employee, date=current_date).first()
                leave = Leave.objects.filter(employee=employee, date=current_date).first()

                status = "Absent"
                check_in = None
                check_out = None
                total_hours = 0

                if leave:
                    status = leave.leave_type
                elif record:
                    check_in = record.check_in
                    check_out = record.check_out
                    status = record.status
                    total_hours = record.total_hours or 0

                results.append({
                    "employee": employee,
                    "date": current_date,
                    "check_in": check_in,
                    "check_out": check_out,
                    "total_hours": total_hours,
                    "status": status,
                })
            current_date += timedelta(days=1)

    return render(request, "hr/attendance/generate_attendance.html", {
        "employees": employees,
        "results": results,
        "from_date": from_date_str,
        "to_date": to_date_str,
        "selected_employee": employee_id,
        "has_searched": has_searched,
    })

    # Add this new function
def attendance_upload_delete(request, pk):
    upload = get_object_or_404(AttendanceUpload, pk=pk)
    # This deletes the database record. 
    # Note: If you want to delete the actual file from your media folder too, 
    # use: upload.file.delete()
    upload.delete()
    messages.success(request, "Upload history record deleted.")
    return redirect("hr:attendance_upload")

# -------------------------
# ATTENDANCE SCHEDULE CRUD
# -------------------------

# 2️⃣ Attendance Schedules List
def schedule_list(request):
    schedules = AttendanceSchedule.objects.all()
    return render(request, "hr/attendance/schedule_list.html", {"schedules": schedules})

# 2️⃣ Schedule Create
def schedule_create(request):
    if request.method == "POST":
        form = AttendanceScheduleForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Attendance schedule created successfully!")
            return redirect("hr:attendance_schedule_list")
        else:
            messages.error(request, "Failed to create schedule. Please correct the errors.")
    else:
        form = AttendanceScheduleForm()

    return render(request, "hr/attendance/schedule_create.html", {
        "form": form,
        "title": "Create Attendance Schedule"
    })

# 3️⃣ Schedule Edit
def schedule_edit(request, schedule_id): # This must match <int:schedule_id> from urls.py
    schedule = get_object_or_404(AttendanceSchedule, id=schedule_id)
    
    if request.method == "POST":
        form = AttendanceScheduleForm(request.POST, instance=schedule)
        if form.is_valid():
            form.save()
            messages.success(request, "Schedule updated!")
            return redirect("hr:attendance_schedule_list")
    else:
        form = AttendanceScheduleForm(instance=schedule)
    
    return render(request, "hr/attendance/schedule_edit.html", {
        "form": form, 
        "schedule": schedule
    })

# 4️⃣ Schedule Delete
def schedule_delete(request, schedule_id):
    schedule = get_object_or_404(AttendanceSchedule, id=schedule_id)
    if request.method == "POST":
        schedule.delete()
        messages.success(request, "Attendance schedule deleted successfully!")
        return redirect("hr:attendance_schedule_list")

    return render(request, "hr/attendance/schedule_delete.html", {"schedule": schedule})



from django.db.models import Q
from .models import Record, AttendanceSchedule

from datetime import datetime, date

def latest_attendance(request):
    schedule = AttendanceSchedule.objects.filter(is_active=True).first()
    shift_start = schedule.check_in_time if schedule else None

    records = Record.objects.select_related("employee").order_by(
        "-date", "-check_in"
    )[:50]

    total_present = 0
    total_late = 0
    total_absent = 0

    attendance_list = []

    for record in records:
        status = "Absent"
        total_hours = 0

        if record.check_in:
            status = "Present"

            if shift_start and record.check_in > shift_start:
                status = "Late"

        if record.check_in and record.check_out:
            d1 = datetime.combine(date.today(), record.check_in)
            d2 = datetime.combine(date.today(), record.check_out)

            if d2 >= d1:
                diff = d2 - d1
                total_hours = round(diff.total_seconds() / 3600, 2)

        if status == "Present":
            total_present += 1
        elif status == "Late":
            total_late += 1
        else:
            total_absent += 1

        attendance_list.append({
            "employee": record.employee,
            "date": record.date,
            "check_in": record.check_in,
            "check_out": record.check_out,
            "status": status,
            "total_hours": total_hours,
        })

    context = {
        "records": attendance_list,
        "total_present": total_present,
        "total_late": total_late,
        "total_absent": total_absent,
    }

    return render(request, "hr/attendance/latest.html", context)


# 6️⃣ Generate Attendance (Logic Placeholder)
# hr/views.py
from datetime import datetime, timedelta # Ensure timedelta is imported here
from datetime import datetime, timedelta
from .models import Record, Employee, Leave, AttendanceSchedule

def generate_attendance(request):
    employees = Employee.objects.filter(status="Active").order_by("name")

    from_date_str = request.GET.get("from_date")
    to_date_str = request.GET.get("to_date")
    employee_id = request.GET.get("employee")

    results = []
    has_searched = False

    if from_date_str and to_date_str:
        has_searched = True

        from_date = datetime.strptime(from_date_str, "%Y-%m-%d").date()
        to_date = datetime.strptime(to_date_str, "%Y-%m-%d").date()

        employee_qs = employees
        if employee_id and employee_id != "all":
            employee_qs = employee_qs.filter(id=employee_id)

        current_date = from_date
        while current_date <= to_date:
            for employee in employee_qs:

                record = Record.objects.filter(
                    employee=employee,
                    date=current_date
                ).first()

                leave = Leave.objects.filter(
                    employee=employee,
                    date=current_date
                ).first()

                check_in = None
                check_out = None
                remarks = ""
                status = "Absent"

                if leave:
                    status = leave.leave_type
                    remarks = "On Leave"

                elif record:
                    check_in = record.check_in
                    check_out = record.check_out
                    status = "Present"

                results.append({
                    "employee": employee,
                    "date": current_date,
                    "check_in": check_in,
                    "check_out": check_out,
                    "status": status,
                    "remarks": remarks,
                })

            current_date += timedelta(days=1)

    return render(request, "hr/attendance/generate_attendance.html", {
        "employees": employees,
        "results": results,
        "from_date": from_date_str,
        "to_date": to_date_str,
        "selected_employee": employee_id,
        "has_searched": has_searched,
    })
    
    
    
# 7️⃣ Add Individual Attendance Record
def attendance_record_create(request):
    if request.method == "POST":
        form = AttendanceRecordForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Attendance record added successfully!")
            return redirect("hr:attendance_latest")

    else:
        form = AttendanceRecordForm()

    return render(request, "hr/attendance/record_form.html", {
        "form": form,
        "title": "Add Attendance Record"
    })




# 8️⃣ Edit Individual Attendance Record
def attendance_record_edit(request, pk):
    record = get_object_or_404(AttendanceRecord, pk=pk)

    if request.method == "POST":
        form = AttendanceRecordForm(request.POST, instance=record)
        if form.is_valid():
            form.save()
            messages.success(request, "Attendance record updated successfully!")
            return redirect("hr:attendance_latest")

    else:
        form = AttendanceRecordForm(instance=record)

    return render(request, "hr/attendance/record_form.html", {
        "form": form,
        "title": "Edit Attendance Record"
    })



from datetime import datetime, date
from .models import AttendanceRecord, AttendanceSchedule

def sync_record_to_attendance_record(record):
    """
    Creates or updates AttendanceRecord from Record
    """
    schedule = AttendanceSchedule.objects.filter(is_active=True).first()
    shift_start = schedule.check_in_time if schedule else None

    status = "Absent"
    remarks = ""

    if record.check_in:
        status = "Present"
        if shift_start and record.check_in > shift_start:
            status = "Late"

    AttendanceRecord.objects.update_or_create(
        employee=record.employee,
        date=record.date,
        defaults={
            "check_in": record.check_in,
            "check_out": record.check_out,
            "status": status,
            "remarks": remarks,
        }
    )




# 9️⃣ Attendance Filtering Report
from django.db.models import Count, Q
from .models import Record, AttendanceRecord, Employee, AttendanceSchedule

def attendance_report(request):
    employees = Employee.objects.all()

    # 🔹 STEP 1: Sync ALL records (no filters)
    all_records = Record.objects.select_related("employee")

    schedule = AttendanceSchedule.objects.filter(is_active=True).first()
    shift_start = schedule.check_in_time if schedule else None

    for record in all_records:
        status = "Absent"

        if record.check_in:
            status = "Present"
            if shift_start and record.check_in > shift_start:
                status = "Late"

        AttendanceRecord.objects.update_or_create(
            employee=record.employee,
            date=record.date,
            defaults={
                "check_in": record.check_in,
                "check_out": record.check_out,
                "status": status,
            }
        )

    # 🔹 STEP 2: Now read AttendanceRecord for report
    results = AttendanceRecord.objects.select_related("employee")

    # Filters
    emp_id = request.GET.get("employee")
    dep = request.GET.get("department")
    des = request.GET.get("designation")
    date_from = request.GET.get("from")
    date_to = request.GET.get("to")

    if emp_id and emp_id != "all":
        results = results.filter(employee_id=emp_id)

    if dep and dep != "all":
        results = results.filter(department=dep)

    if des and des != "all":
        results = results.filter(designation=des)

    if date_from:
        results = results.filter(date__gte=date_from)

    if date_to:
        results = results.filter(date__lte=date_to)

    department_summary = (
        results.values("department")
        .annotate(
            total=Count("id"),
            present=Count("id", filter=Q(status="Present")),
            late=Count("id", filter=Q(status="Late")),
            absent=Count("id", filter=Q(status="Absent")),
        )
    )

    return render(request, "hr/reports/attendance_report.html", {
        "employees": employees,
        "departments": [d[0] for d in DEPARTMENT_CHOICES],
        "designations": [d[0] for d in DESIGNATION_CHOICES],
        "results": results.order_by("-date"),
        "department_summary": department_summary,
    })



def attendance_report_excel(request):
    qs = AttendanceRecord.objects.select_related("employee").order_by("-date")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance Report"

    headers = [
        "Emp ID", "Name", "Department", "Designation",
        "Status", "Check In", "Check Out", "Hours", "Date", "Remarks"
    ]
    ws.append(headers)

    for r in qs:
        ws.append([
            r.employee.emp_id,
            r.employee.name,
            r.employee.department,
            r.employee.designation,
            r.status,
            str(r.check_in),
            str(r.check_out),
            r.total_hours,
            str(r.date),
            r.remarks or "",
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=attendance_report.xlsx"
    wb.save(response)
    return response

from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from django.http import HttpResponse

def attendance_report_pdf(request):
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'inline; filename="attendance_report.pdf"'

    doc = SimpleDocTemplate(
        response,
        pagesize=A4,
        rightMargin=30,
        leftMargin=30,
        topMargin=40,
        bottomMargin=30,
    )

    elements = []
    styles = getSampleStyleSheet()

    # 🔹 TITLE
    elements.append(Paragraph("<b>Attendance Report</b>", styles["Title"]))

    elements.append(Paragraph("<br/>", styles["Normal"]))

    # 🔹 TABLE HEADER
    table_data = [[
        "Emp ID",
        "Name",
        "Department",
        "Designation",
        "Status",
        "Check In",
        "Check Out",
        "Total Hours",
        "Date",
    ]]

    # 🔹 FETCH DATA
    records = AttendanceRecord.objects.select_related("employee").order_by("-date")

    for r in records:
        table_data.append([
            r.employee.emp_id,
            r.employee.name,
            r.department,
            r.designation,
            r.status,
            r.check_in.strftime("%H:%M") if r.check_in else "-",
            r.check_out.strftime("%H:%M") if r.check_out else "-",
            f"{r.total_hours} hrs" if r.total_hours else "0",
            r.date.strftime("%Y-%m-%d"),
        ])

    # 🔹 CREATE TABLE
    table = Table(table_data, repeatRows=1)

    table.setStyle(TableStyle([
        # Header styling
        ("BACKGROUND", (0, 0), (-1, 0), colors.darkblue),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 10),

        # Cell styling
        ("FONTSIZE", (0, 1), (-1, -1), 9),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),

        # Grid
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),

        # Padding
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))

    elements.append(table)
    doc.build(elements)

    return response



from django.db.models import Count, Q
from django.utils import timezone
from .models import AttendanceRecord

def monthly_summary_report(request):
    month_list = [
        (1, "January"), (2, "February"), (3, "March"), (4, "April"),
        (5, "May"), (6, "June"), (7, "July"), (8, "August"),
        (9, "September"), (10, "October"), (11, "November"), (12, "December"),
    ]

    results = None
    month = request.GET.get("month")
    year = request.GET.get("year")

    selected_month = int(month) if month and month.isdigit() else None
    selected_year = int(year) if year and year.isdigit() else timezone.now().year

    if selected_month and selected_year:
        results = (
            AttendanceRecord.objects
            .filter(
                date__month=selected_month,
                date__year=selected_year
            )
            .values(
                "employee_id",
                "employee__emp_id",
                "employee__name",
                "department",
                "designation",
            )
            .annotate(
                present=Count("id", filter=Q(status="Present")),
                late=Count("id", filter=Q(status="Late")),
                absent=Count("id", filter=Q(status="Absent")),
                leave_full=Count("id", filter=Q(status="Full Leave")),
                leave_short=Count("id", filter=Q(status="Short Leave")),
                on_duty=Count("id", filter=Q(status="On Duty")),
                total_days=Count("id"),
            )
            .order_by("employee__emp_id")
        )

        # Optional: Calculate payable days in Python
        for row in results:
            row["payable_days"] = (
                row["present"]
                + row["late"]
                + row["on_duty"]
                + row["leave_full"]
                + row["leave_short"]
            )

    context = {
        "results": results,
        "month_list": month_list,
        "selected_month": selected_month,
        "selected_year": selected_year,
    }

    return render(request, "hr/reports/monthly_summary.html", context)

from django.db.models import F, Count, Q
from django.utils import timezone

def department_stats_report(request):
    month_list = [
        (1, "January"), (2, "February"), (3, "March"), (4, "April"),
        (5, "May"), (6, "June"), (7, "July"), (8, "August"),
        (9, "September"), (10, "October"), (11, "November"), (12, "December"),
    ]

    results = None
    month = request.GET.get("month")
    year = request.GET.get("year")

    selected_month = int(month) if month and month.isdigit() else None
    selected_year = int(year) if year and year.isdigit() else timezone.now().year

    if selected_month and selected_year:
        # We rename the group field to 'dept_code' to avoid conflict with the model's 'department' field
        department_data = AttendanceRecord.objects.filter(
            date__month=selected_month,
            date__year=selected_year
        ).values(
            dept_code=F('employee__department') 
        ).annotate(
            total_records=Count("id"), 
            present_days=Count("id", filter=Q(status="Present")) 
        )

        results = list(department_data)
        
        # Prepare the mapping for display names
        dept_name_map = dict(DEPARTMENT_CHOICES)

        for r in results:
            if r["total_records"] > 0:
                r["attendance_rate"] = round((r["present_days"] / r["total_records"]) * 100, 2)
            else:
                r["attendance_rate"] = 0
            
            # Map the code to the readable name (e.g., 'HR' -> 'Human Resources')
            r["department_name"] = dept_name_map.get(r["dept_code"], r["dept_code"])

    context = {
        "results": results,
        "month_list": month_list,
        "selected_month": selected_month,
        "selected_year": selected_year
    }

    return render(request, "hr/reports/department_stats.html", context)